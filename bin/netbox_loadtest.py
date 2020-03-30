#!/usr/bin/env python3

"""
Test the netbox API by running a set of scenarios at it designed to show how netbox responds to heavy load.

1. Allocate addresses using get next free logic, then deallocate those addresses.
2. Randomly grab addresses, then deallocate them in that order.
3. Use get next free address logic in a heavily fragmented prefix, then deallocate only those addresses.

After running each of these 3 tests, results are combined and a more general report is generated.
"""

import argparse
import ipaddress
import json
import queue
import random
import threading
import time

import requests
from openpyxl import Workbook

parser = argparse.ArgumentParser(description='Test the Netbox API.')
parser.add_argument('parent_prefix', type=str, help='the prefix the worker should pull the child prefix from')
parser.add_argument('prefix_length', type=int, help='the size of the prefix to carve out')
parser.add_argument('workers', type=int, help='number of workers concurrenting working')
parser.add_argument('url', type=str, help='FQDN of netbox')
parser.add_argument('token', type=str, help='Auth token for netbox API')

args = parser.parse_args()

URL = f'http://{ args.url }/api'
HEADERS = {
    'accept': 'application/json',
    'Content-Type': 'application/json',
    'Authorization': f'Token { args.token }'
}

session = requests.Session()
session.headers = HEADERS


def get_prefix(prefix: str) -> dict:
    """Get a prefix object from the API."""

    url = '{}/ipam/prefixes/?prefix={}'.format(URL, prefix.replace('/', '%2F'))
    result = session.get(url)

    return result.json()['results'][0]


def delete_prefix(prefix: dict) -> int:
    """Call API to delete the prefix."""

    id = prefix['id']
    result = session.delete(f'{ URL }/ipam/prefixes/{ id }/')

    return result


def carve_new_prefix(parent_prefix: str, prefix_length: int) -> dict:
    """Call API to create a new prefix object from parent prefix object."""

    data = {"prefix_length": prefix_length, 'description': 'test_netbox created this prefix'}
    id = get_prefix(parent_prefix)['id']
    result = session.post(f'{ URL }/ipam/prefixes/{ id }/available-prefixes/', data=json.dumps(data))

    return result.json()


def get_and_reserve_next_free_address(prefix: dict) -> dict:
    """Call API to allocate a new address from prefix."""

    data = {'description': 'test_netbox allocated this address'}
    id = prefix['id']
    result = session.post(f'{ URL }/ipam/prefixes/{ id }/available-ips/', data=json.dumps(data))

    if result.status_code == 204:
        raise RuntimeError

    return result.json()


def deallocate_address(address: dict) -> bool:
    """Remove assignment data from an address."""

    id = address['id']
    response = session.delete(f'{ URL }/ipam/ip-addresses/{ id }/')

    return True if response.status_code == 204 else False


def reserve_address(address: str, description="allocated as part of a test") -> dict:
    """Assign data to fields of an address to claim it."""

    data = {"address": address, "description": description}
    response = session.post(f'{ URL }/ipam/ip-addresses/', data=json.dumps(data))

    return response.json()


def test_get_next_free_address(prefix: dict) -> dict:
    """Use the get next free logic to allocate every address in prefix then deallocate."""

    report = {'prefix': prefix['prefix'], 'allocate': {'data': {}}, 'deallocate': {'data': {}}}
    addresses_assigned = []

    while True:
        try:
            start = time.time()
            address = get_and_reserve_next_free_address(prefix)
            _address = address['address'].split('/', 1)[0]
            report['allocate']['data'][_address] = time.time() - start
            addresses_assigned.append(address)
        except RuntimeError:
            break

    for address in addresses_assigned:
        _address = address['address'].split('/', 1)[0]
        start = time.time()
        if deallocate_address(address):
            report['deallocate']['data'][_address] = time.time() - start

    return report


def test_get_next_free_address_fragmented(prefix: dict) -> dict:
    """Use the get next free logic to allocate every other address in prefix then deallocate them."""

    report = {'prefix': prefix['prefix'], 'allocate': {'data': {}}, 'deallocate': {'data': {}}}
    addresses_assigned = []
    prefix_obj = ipaddress.IPv4Network(prefix['prefix'])
    fragmentated_addresses = []

    # create the fragmentation
    for address_obj in prefix_obj.hosts():
        if int(address_obj) % 2:
            address = reserve_address(str(address_obj), 'fragmentation for a test')
            fragmentated_addresses.append((address_obj, address))

    while True:
        try:
            start = time.time()
            address = get_and_reserve_next_free_address(prefix)
            _address = address['address'].split('/', 1)[0]
            report['allocate']['data'][_address] = time.time() - start
            addresses_assigned.append(address)
        except RuntimeError:
            break

    for address in addresses_assigned:
        _address = address['address'].split('/', 1)[0]
        start = time.time()
        deallocate_address(address)
        report['deallocate']['data'][_address] = time.time() - start

    # clean up fragmentation
    for address_obj, address in fragmentated_addresses:
        if int(address_obj) % 2:
            deallocate_address(address)

    return report


def test_scattered_assignments(prefix: dict) -> dict:
    """Execute a non-linear pattern of allocating addresses and then deallocating them."""

    addresses_to_unassign = []
    report = {'prefix': prefix['prefix'], 'allocate': {'data': {}}, 'deallocate': {'data': {}}}

    addresses_to_assign = [str(address) for address in ipaddress.ip_network(prefix['prefix']).hosts()]
    random.shuffle(addresses_to_assign)

    for address in addresses_to_assign:
        start = time.time()
        _address = reserve_address(address)
        if _address:
            addresses_to_unassign.append(_address)
            report['allocate']['data'][address] = time.time() - start

    for address in addresses_to_unassign:
        start = time.time()
        _address = address['address'].split('/', 1)[0]
        if deallocate_address(address):
            report['deallocate']['data'][_address] = time.time() - start

    return report


def worker(prefix: dict):
    """Execute all 3 scenarios against prefix then save the report."""

    print('    testing with {}'.format(prefix['prefix']))
    report = {}
    start = time.time()

    report['test_get_next_free_address'] = test_get_next_free_address(prefix)
    report['test_get_next_free_address_fragmented'] = test_get_next_free_address_fragmented(prefix)
    report['test_scattered_assignments'] = test_scattered_assignments(prefix)

    report['total_duration'] = time.time() - start

    report_queue.put(report)
    print('    finished with {}'.format(prefix['prefix']))


def colnum_string(n):
    """credit to https://stackoverflow.com/questions/23861680/convert-spreadsheet-number-to-column-letter#"""
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def add_worker_data_to_sheet(worker_data: dict, sheet: object):
    """Add all the worker reports in worker_data to the workbook.

    We create a tab for every level of concurrency from 1 - N.
    Each tab contains 6 datsets (3 tests with 2 phases each) worth of raw data, avg, stdv and total.
    Hopefully the raw data will make sure useful data visualization which is not included in this code.
    """

    for worker_id, worker_report in enumerate(worker_data.values()):
        footer_row = len(worker_report['test_get_next_free_address']['allocate']['data']) + 6
        sheet[f'B{ footer_row }'] = 'mean'
        sheet[f'B{ footer_row + 1 }'] = 'stdev'
        sheet[f'B{ footer_row + 2 }'] = 'total'

        for test_name, test_id in {
            'test_get_next_free_address': 1,
            'test_get_next_free_address_fragmented': 2,
            'test_scattered_assignments': 3
        }.items():
            row = 1
            column = (worker_id * 6) + (test_id * 2) + 1
            lcolumn = colnum_string(column)  # get the letter value for y-coord
            lcolumn2 = colnum_string(column + 1)  # get the letter value for y-coord

            sheet.cell(row=1, column=column).value = f'worker { worker_id + 1 }'
            sheet.cell(row=1, column=column + 1).value = f'worker { worker_id + 1}'
            sheet.cell(row=2, column=column).value = test_name
            sheet.cell(row=2, column=column + 1).value = test_name
            sheet.cell(row=3, column=column).value = 'allocate'
            sheet.cell(row=3, column=column + 1).value = 'deallocate'

            for id, value in enumerate(worker_report[test_name]['allocate']['data'].values()):
                sheet.cell(row=id + 4, column=column).value = value
                if (id + 4) > row:
                    row = id + 4

            sheet.cell(row=footer_row, column=column).value = f'=AVERAGE({lcolumn}4:{lcolumn}{row})'
            sheet.cell(row=footer_row + 1, column=column).value = f'=STDEV({lcolumn}4:{lcolumn}{row})'
            sheet.cell(row=footer_row + 2, column=column).value = f'=SUM({lcolumn}4:{lcolumn}{row})'

            for id, key in enumerate(worker_report[test_name]['deallocate']['data'].keys()):
                sheet.cell(row=id + 4, column=column + 1).value = worker_report[test_name]['deallocate']['data'][key]
                if (id + 4) > row:
                    row = id + 4

            sheet.cell(row=footer_row, column=column + 1).value = f'=AVERAGE({lcolumn2}4:{lcolumn2}{row})'
            sheet.cell(row=footer_row + 1, column=column + 1).value = f'=STDEV({lcolumn2}4:{lcolumn2}{row})'
            sheet.cell(row=footer_row + 2, column=column + 1).value = f'=SUM({lcolumn2}4:{lcolumn2}{row})'


report_queue = queue.Queue()

""" Spawn some worker threads and load test the NetBox API and then make an excel spreadsheet about it."""

if __name__ == "__main__":
    workbook = Workbook()
    worker_data = {}

    for worker_max in range(1, args.workers + 1):
        threads = []
        print(f'starting the { worker_max } worker scenario')

        for worker_id in range(1, worker_max + 1):
            prefix = carve_new_prefix(args.parent_prefix, args.prefix_length)
            _prefix = prefix['prefix']
            print(f'  starting worker thread { worker_id } of { worker_max } with { _prefix }')
            thread = threading.Thread(target=worker, args=(prefix,))
            thread.start()
            threads.append((thread, prefix))

        for thread, prefix in threads:
            thread.join()
            delete_prefix(prefix)
            worker_data[prefix['prefix']] = report_queue.get()

        sheet = workbook.create_sheet(f'{ worker_max } workers')
        add_worker_data_to_sheet(worker_data, sheet)

    workbook.save(filename='netbox_load_test_report_{}.xlsx'.format(args.parent_prefix.replace('/', '_')))
